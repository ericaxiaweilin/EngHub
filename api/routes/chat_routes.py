"""
AI Assistant chat routes（支持 Tool Calling）。

代理到 litellm 网关 (OpenAI 兼容 /v1/chat/completions)。
- 纯问答：直接转发对话。
- 操作型：通过 function-calling 让模型调用 MES 工具（查工单/建工单/报工/查库存等），
  后端执行工具并把结果回传给模型生成最终回复，同时把"已执行的操作"返回给前端展示。
所有连接参数通过环境变量配置，未配置或网关不可达时返回友好降级回复，保证前端可用。
"""

from __future__ import annotations

import asyncio
import base64
import json
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.parse import quote

import httpx
from fastapi import APIRouter, Depends, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from database.db_config import get_db
from database.models import FileRecord, User
from core.auth.security import get_current_user
from api.services.chat_tools_service import (
    TOOL_DEFINITIONS, TOOL_LABELS, WRITE_TOOLS, SIM_TOOLS, execute_tool,
)
from api.services.quick_command_service import (
    build_agent_system_prompt, record_agent_dispatch,
)

router = APIRouter(prefix="/api/v1/chat", tags=["ai-assistant"])

# --- 模型底座接入配置 ---
GATEWAY_URL = os.getenv("LLM_GATEWAY_URL", "http://host.docker.internal:14040").rstrip("/")
API_KEY = os.getenv("LLM_API_KEY", "")
REQUEST_TIMEOUT = float(os.getenv("LLM_TIMEOUT", "60"))
MAX_TOOL_ROUNDS = int(os.getenv("LLM_MAX_TOOL_ROUNDS", "5"))
MODEL_STACK_CONTROL_PLANE_URL = os.getenv("MODEL_STACK_CONTROL_PLANE_URL", "").rstrip("/")
MODEL_STACK_CHAT_TASK_ID = os.getenv("MODEL_STACK_CHAT_TASK_ID", "").strip()
MODEL_STACK_VISION_TASK_ID = os.getenv("MODEL_STACK_VISION_TASK_ID", "").strip()
MODEL_STACK_ROUTE_TIMEOUT = float(os.getenv("MODEL_STACK_ROUTE_TIMEOUT", "5"))

SYSTEM_PROMPT = (
    "你是 EngHub MES 制造执行系统的智能助手，可以直接操作系统完成用户的请求。"
    "你熟悉生产工单、报工、检验、不良品、库存、生产计划(MRP)、"
    "工位/工艺/设备、员工技能矩阵以及合规仿真引擎(Sim-ERP)等模块。\n"
    "重要：当用户要求查询数据或执行操作（如查工单、建工单、报工、查库存、查不良品、查设备、下达工单、"
    "完工/暂停/拆分工单、查工艺路线、查技能矩阵、运行合规仿真、查仿真审计记录等）时，"
    "你必须调用对应的工具(tool)来获取真实数据或完成操作，不要凭空编造数据。"
    "写操作（创建工单/下达工单/报工/完工等）执行后，请向用户确认操作结果。\n"
    "【工作流优先】当用户请求复合任务（如'帮我复盘今天生产'、'质量异常分诊'、'全面合规检查'、'建一个工单并下达'）时，"
    "优先调用 run_workflow 工具运行预置工作流（生产日度复盘/质量异常分诊/全面合规检查/一键建单下达），"
    "而不是逐个调用单步工具。\n"
    "【多模态附件】用户可能上传图片或文件。若收到图片，请结合图片内容回答（如识别设备/工件/缺陷/图纸/仪表盘），"
    "描述你看到的内容并给出专业判断；若用户要求“识别/OCR/提取文字”，请完整提取图片中所有文字内容（保留原始结构与格式）；"
    "若收到文件，基于文字与附件信息回答。\n"
    "【严禁推诿】绝对不要回答“建议你进入XX看板/日报中心/实时看板查看”、"
    "“具体数值需结合你的实时数据源/PLC采集”这类把用户打发走的话。"
    "你能直接读到真实数据库，必须立即调用工具取数并以表格/清单形式呈现给用户。\n"
    "【预警情报中枢】你不仅是查询助手，更是预警情报审查员。当系统产生被动预警（安灯工单/质量缺陷/设备故障/工单超时）时，"
    "你会自动进行初步审查（严重度/根因/建议/分派）。用户可随时问你“有什么预警”“预警简报”获取当前态势，"
    "也可以说“巡检”让你主动扫描异常。只有工具返回了对应证据时，才能给出根因、处置建议和分派对象。\n"
    "【流程知识库】系统内置了完整的流程知识：工单全生命周期（8阶段：创建→下达→派工→执行→报工→质检→完工→入库）、"
    "6大职位标准作业流程(操作员/品检员/设备工程师/PMC计划员/生产主管/仓管员)、各环节RACI责任矩阵。"
    "用户问流程/职责/该找谁类问题时，调用 query_process_knowledge 工具获取标准答案。\n"
    "【任务中心】工业场景很多任务无法一次完成（等物料/等审批/等设备恢复/等供应商）。"
    "当用户交代的事情当前无法闭环、或用户说'跟进一下''盯着这个''挂起来''到时候提醒我'时，"
    "调用 create_followup_task 把任务挂入任务中心，系统会按频率（默认2小时，用户可指定）定期自动跟进并推送通知；"
    "挂账成功后告知用户可在「任务中心」页面查看进度。不要把能立即完成的查询/操作挂账。\n"
    "【结果解读】工具与工作流返回的 JSON 只是你的输入素材，绝对不能原样贴给用户。""你必须把它翻译成人话：先给结论，再用表格或短清单列关键数据，最后说明下一步。""多步工作流要逐步说明每一步查到了什么，而不是复述 steps 数组。\n"    "【回答边界】工具结果未提供的信息不得猜测，不得擅自补充故障、缺料、同步异常等可能原因。"
    "最终回答只输出面向用户的结论，禁止输出 <think>、推理过程、内部分析或工具选择过程。\n"
    "请用简洁专业的中文回答制造与车间管理相关问题。"
)

TOOL_RESULT_GROUNDING = (
    "以下 JSON 是本次回答唯一可用的业务事实。只陈述 JSON 明确提供的数据，"
    "不得补充可能原因、假设、示例编号、风险结论或系统未执行的动作。"
    "数值为 0 或列表为空时，只说明当前返回无记录，不推测原因。"
)
FINAL_GROUNDING_PROMPT = (
    "最终答复必须逐项对应前面的工具 JSON。禁止添加 JSON 中不存在的状态、"
    "原因、风险、预警、示例、建议或已执行动作；不要用常识补全缺失字段。"
)


class ChatMessage(BaseModel):
    role: str
    content: str


class Attachment(BaseModel):
    """随消息提交的附件引用（前端先调 /files/upload 拿 file_id，再随消息提交）。"""
    file_id: str
    kind: Optional[str] = None  # image / file，缺省时按 content_type 推断


class ChatRequest(BaseModel):
    messages: List[ChatMessage]
    temperature: float = 0.3
    enable_tools: bool = True  # 是否启用工具调用
    attachments: List[Attachment] = Field(default_factory=list)  # 本轮用户消息附带的附件
    agent_key: Optional[str] = None  # 指定调度的智能体（空=自动，由模型自行选择工具）


class ToolAction(BaseModel):
    """一次工具执行记录，供前端展示'AI 已执行的操作'。"""
    tool: str
    label: str
    arguments: Dict[str, Any] = Field(default_factory=dict)
    result: Dict[str, Any] = Field(default_factory=dict)
    is_write: bool = False
    is_sim: bool = False
    success: bool = True


class ChatResponse(BaseModel):
    reply: str
    model: str
    degraded: bool = False
    actions: List[ToolAction] = Field(default_factory=list)


@router.get("/health")
async def chat_health():
    """返回模型底座任务路由与网关连通性状态。"""
    configured = bool(
        GATEWAY_URL
        and MODEL_STACK_CONTROL_PLANE_URL
        and MODEL_STACK_CHAT_TASK_ID
    )
    reachable = False
    detail = "model-stack configuration incomplete"
    if configured:
        try:
            route = await _resolve_model_route(MODEL_STACK_CHAT_TASK_ID)
            gateway_headers = {"Authorization": f"Bearer {API_KEY}"} if API_KEY else {}
            async with httpx.AsyncClient(timeout=5.0) as client:
                gateway_resp = await client.get(
                    f"{GATEWAY_URL}/v1/models", headers=gateway_headers,
                )
            reachable = gateway_resp.status_code < 400
            detail = (
                f"control-plane=ready, gateway_models={gateway_resp.status_code}, "
                f"route={route['task_id']}"
            )
        except Exception as exc:  # noqa: BLE001
            detail = f"unreachable: {type(exc).__name__}"
    return {
        "configured": configured,
        "reachable": reachable,
        "model": MODEL_STACK_CHAT_TASK_ID,
        "gateway": GATEWAY_URL,
        "control_plane": MODEL_STACK_CONTROL_PLANE_URL,
        "detail": detail,
    }


@router.get("/tools")
async def chat_tools():
    """返回当前可用的 MES 工具清单与工作流清单（供前端展示能力/快捷指令）。"""
    from api.services.workflow_service import list_workflows  # 懒加载，避免循环导入
    return {
        "tools": [
            {
                "name": t["function"]["name"],
                "label": TOOL_LABELS.get(t["function"]["name"], t["function"]["name"]),
                "description": t["function"]["description"],
                "is_write": t["function"]["name"] in WRITE_TOOLS,
                "is_sim": t["function"]["name"] in SIM_TOOLS,
            }
            for t in TOOL_DEFINITIONS
        ],
        "workflows": list_workflows(),
    }


async def _resolve_model_route(
    task_id: str,
    *,
    prompt_tokens: int = 1000,
    max_completion_tokens: int = 1024,
) -> Dict[str, Any]:
    """向模型底座申请任务路由；业务侧不维护模型候选或回退链。"""
    if not MODEL_STACK_CONTROL_PLANE_URL or not task_id:
        raise RuntimeError("model-stack task routing is not configured")

    url = (
        f"{MODEL_STACK_CONTROL_PLANE_URL}/api/model-management/"
        f"business-tasks/{quote(task_id, safe='')}/route-request"
    )
    params = {
        "prompt_tokens": max(0, int(prompt_tokens)),
        "max_completion_tokens": max(0, int(max_completion_tokens)),
        "require_deployed": "true",
    }
    async with httpx.AsyncClient(timeout=MODEL_STACK_ROUTE_TIMEOUT) as client:
        resp, manifest_resp = await asyncio.gather(
            client.get(url, params=params),
            client.get(
                f"{MODEL_STACK_CONTROL_PLANE_URL}/api/model-management/providers/deployed"
            ),
        )
    resp.raise_for_status()
    manifest_resp.raise_for_status()

    envelope = resp.json()
    route = envelope.get("route_request") if isinstance(envelope, dict) else None
    providers = route.get("providers") if isinstance(route, dict) else None
    provider = str(providers[0] if providers else "").strip()
    if not provider:
        raise RuntimeError(f"model-stack returned no deployed route for {task_id}")

    manifest = manifest_resp.json()
    provider_rows = manifest.get("providers") if isinstance(manifest, dict) else None
    provider_row = next(
        (
            row for row in (provider_rows or [])
            if isinstance(row, dict)
            and str(row.get("provider") or row.get("key") or "").strip() == provider
        ),
        None,
    )
    gateway_model = str(
        (provider_row or {}).get("target_model")
        or (provider_row or {}).get("model")
        or ""
    ).strip()
    if not gateway_model:
        raise RuntimeError(f"model-stack returned no execution target for {provider}")

    runtime_policy = route.get("runtime_policy") or {}
    timeout_ms = (
        route.get("request_timeout_ms")
        or runtime_policy.get("request_timeout_ms")
        or int(REQUEST_TIMEOUT * 1000)
    )
    completion_limit = (
        route.get("max_completion_tokens")
        or runtime_policy.get("max_completion_tokens")
        or max_completion_tokens
    )
    return {
        "task_id": str(route.get("dispatch_scenario") or task_id),
        "provider": provider,
        "gateway_model": gateway_model,
        "request_timeout": max(1.0, float(timeout_ms) / 1000),
        "max_completion_tokens": max(1, int(completion_limit)),
    }


async def _call_llm(
    payload: Dict[str, Any],
    *,
    request_timeout: Optional[float] = None,
) -> httpx.Response:
    """通过模型底座网关调用控制面下发的 provider。"""
    headers = {
        "Content-Type": "application/json",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
    }
    if API_KEY:
        headers["Authorization"] = f"Bearer {API_KEY}"
    async with httpx.AsyncClient(timeout=request_timeout or REQUEST_TIMEOUT) as client:
        return await client.post(
            f"{GATEWAY_URL}/v1/chat/completions",
            json=payload,
            headers=headers,
        )


def _clean_model_reply(content: str) -> str:
    """清除模型协议中误混入 content 的推理区块，不改变最终答案语义."""
    reply = (content or "").strip()
    return re.sub(
        r"<think>.*?</think>",
        "",
        reply,
        flags=re.DOTALL | re.IGNORECASE,
    ).strip()


def _grounded_tool_result(result: Dict[str, Any]) -> str:
    return (
        f"{TOOL_RESULT_GROUNDING}\n"
        f"{json.dumps(result, ensure_ascii=False, default=str)}"
    )


async def _verify_grounded_reply(
    reply: str,
    actions: List[ToolAction],
    route: Dict[str, Any],
) -> str:
    """让模型按工具事实审校草稿；业务侧不参与语义改写。"""
    if not actions:
        return reply
    facts = [
        {"tool": action.tool, "result": action.result}
        for action in actions
    ]
    verification_payload = {
        "model": route["gateway_model"],
        "messages": [
            {
                "role": "system",
                "content": (
                    "你是事实审校器。仅保留能从工具 JSON 逐项验证的陈述，"
                    "删除所有原因猜测、风险推断、预警、建议、示例和未执行动作。"
                    "保持中文自然表达，只输出修订后的最终答复。"
                ),
            },
            {
                "role": "user",
                "content": (
                    f"工具事实：\n{json.dumps(facts, ensure_ascii=False, default=str)}"
                    f"\n\n待审校草稿：\n{reply}"
                ),
            },
        ],
        "temperature": 0,
        "max_tokens": route["max_completion_tokens"],
    }
    resp = await _call_llm(
        verification_payload,
        request_timeout=route["request_timeout"],
    )
    if resp.status_code >= 400:
        return reply
    data = resp.json()
    content = (
        data.get("choices", [{}])[0]
        .get("message", {})
        .get("content", "")
    )
    return _clean_model_reply(content) or reply


async def _load_attachment_records(
    db: AsyncSession, attachments: List[Attachment], user: User,
) -> List[FileRecord]:
    """按 file_id 加载附件记录（做工厂隔离：普通用户不可引用其他工厂文件）。"""
    records: List[FileRecord] = []
    for att in attachments:
        rec = (await db.execute(
            select(FileRecord).where(FileRecord.id == att.file_id)
        )).scalar_one_or_none()
        if not rec:
            continue
        if not user.is_superuser and rec.factory_id and user.factory_id \
                and rec.factory_id != user.factory_id:
            continue  # 跨工厂附件直接忽略，避免越权
        records.append(rec)
    return records


def _is_image_record(rec: FileRecord) -> bool:
    return (rec.content_type or "").startswith("image/")


def _build_multimodal_content(text: str, image_records: List[FileRecord]) -> Any:
    """构造 OpenAI 多模态 content：文本 + 图片(base64 data URL)。

    图片实体从 UPLOAD_DIR 落盘文件读取并转 base64；实体缺失则跳过。
    无可用图片时退化为纯文本字符串。"""
    parts: List[Dict[str, Any]] = [{"type": "text", "text": text}]
    for rec in image_records:
        try:
            data = Path(rec.storage_path).read_bytes()
        except OSError:
            continue
        b64 = base64.b64encode(data).decode("ascii")
        ctype = rec.content_type or "image/png"
        parts.append({
            "type": "image_url",
            "image_url": {"url": f"data:{ctype};base64,{b64}"},
        })
    if len(parts) == 1:
        return text
    return parts


# ---- Excel/CSV 附件解析（让模型真正"读到"表格内容，而非只知道文件名） ----

_SPREADSHEET_CONTENT_TYPES = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/vnd.ms-excel",  # .xls
    "text/csv",
    "application/csv",
)


def _is_spreadsheet_record(rec: FileRecord) -> bool:
    """判断附件是否为 Excel/CSV 表格文件（按扩展名 + content_type 双重识别）。"""
    fn = (rec.filename or "").lower()
    ct = (rec.content_type or "").lower()
    return fn.endswith((".xlsx", ".xlsm", ".xls", ".csv")) or ct in _SPREADSHEET_CONTENT_TYPES


def _parse_spreadsheet_record(
    rec: FileRecord, max_rows: int = 100, max_cols: int = 20,
) -> Optional[Dict[str, Any]]:
    """解析 Excel/CSV 附件 → 结构化表格数据（与前端 TableData 结构一致）。

    返回 {title, columns:[{key,label}], rows:[{...}]}；解析失败返回 None。
    首行视作表头；截断 max_rows/max_cols 避免超大文件撑爆模型上下文。"""
    try:
        path = Path(rec.storage_path)
        if not path.is_file():
            return None
        fn = (rec.filename or "").lower()
        ct = (rec.content_type or "").lower()
        grid: List[List[Any]] = []
        if fn.endswith(".csv") or ct in ("text/csv", "application/csv"):
            import csv as _csv
            with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
                for row in _csv.reader(f):
                    grid.append(list(row))
                    if len(grid) >= max_rows + 1:
                        break
        elif fn.endswith((".xlsx", ".xlsm")) or "spreadsheetml" in ct:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                ws = wb.active
                for row in ws.iter_rows(values_only=True, max_row=max_rows + 1):
                    grid.append(list(row))
            finally:
                wb.close()
        else:
            return None  # .xls 等暂不支持的格式 → 退化为普通文件提示

        # 去掉全空行
        grid = [r for r in grid if any(c is not None and str(c).strip() != "" for c in r)]
        if not grid:
            return None

        header = [
            str(c).strip() if c is not None and str(c).strip() else f"列{i + 1}"
            for i, c in enumerate(grid[0][:max_cols])
        ]
        columns = [{"key": f"c{i}", "label": h} for i, h in enumerate(header)]
        rows: List[Dict[str, Any]] = []
        for raw in grid[1:]:
            rows.append({
                f"c{i}": ("" if i >= len(raw) or raw[i] is None else str(raw[i]))
                for i in range(len(header))
            })
        return {"title": rec.filename or "上传表格", "columns": columns, "rows": rows}
    except Exception:  # noqa: BLE001
        return None


def _spreadsheet_to_summary(table: Dict[str, Any], sample_rows: int = 30) -> str:
    """智能摘要：表头 + 统计 + 样本行，替代全量 Markdown dump。

    小表（≤ sample_rows 行）给出全部行以保证 AI 分析精确；大表仅给统计+前 N 行样本省 token，
    完整数据由前端 Univer 在线表格渲染。明确告知模型数据是否完整，防止其编造表中不存在的行/值。"""
    cols = table["columns"]
    rows = table["rows"]
    total_rows = len(rows)
    col_labels = [c["label"] for c in cols]

    # --- 推断列类型 + 统计 ---
    col_stats: List[str] = []
    for c in cols:
        key, label = c["key"], c["label"]
        values = [r.get(key, "") for r in rows if r.get(key, "") != ""]
        if not values:
            col_stats.append(f"  - {label}: 全空")
            continue
        # 尝试数值推断
        nums = []
        for v in values:
            try:
                nums.append(float(str(v).replace(",", "")))
            except (ValueError, TypeError):
                break
        if nums and len(nums) == len(values):
            col_stats.append(
                f"  - {label} [数值]: min={min(nums):.2f}, max={max(nums):.2f}, "
                f"avg={sum(nums)/len(nums):.2f}, 非空{len(nums)}条"
            )
        else:
            unique = set(str(v) for v in values)
            top = list(unique)[:5]
            col_stats.append(
                f"  - {label} [文本]: {len(unique)}种取值"
                + (f", 如: {', '.join(top)}" if len(unique) <= 20 else f", 前5: {', '.join(top)}")
            )

    # --- 样本行：小表给全部行（保证精确），大表截断（省 token） ---
    shown = rows[:sample_rows]
    is_full = total_rows <= sample_rows
    sample_lines = []
    if shown:
        sample_lines.append("| " + " | ".join(col_labels) + " |")
        sample_lines.append("|" + "|".join(["---"] * len(cols)) + "|")
        for r in shown:
            sample_lines.append("| " + " | ".join(str(r.get(c["key"], "")) for c in cols) + " |")

    parts = [
        f"共 {total_rows} 行 × {len(cols)} 列",
        "列信息：\n" + "\n".join(col_stats),
    ]
    if sample_lines:
        header = (
            "全部数据如下（请严格基于这些真实数据分析，禁止编造表中不存在的行或数值）："
            if is_full else
            f"前 {sample_rows} 行样本（共 {total_rows} 行，其余见用户在线表格）："
        )
        parts.append(header + "\n" + "\n".join(sample_lines))
    if is_full:
        parts.append("（以上即该表格的全部行；分析时不得新增或修改任何数据）")
    else:
        parts.append("（完整数据已在用户的在线表格中展示，用户可直接查看/筛选/编辑）")
    return "\n".join(parts)


def _attachment_text_note(records: List[FileRecord]) -> str:
    """附件文字摘要。

    - Excel/CSV：智能摘要（表头+统计+样本），完整数据由 Univer 在线表格渲染；
    - 图片/其他文件：仅告知文件名/类型/大小（用于不支持 vision 时的优雅降级）。"""
    if not records:
        return ""
    lines = ["\n\n【用户本次上传的附件（已存入系统文件库）】"]
    for rec in records:
        if _is_spreadsheet_record(rec):
            table = _parse_spreadsheet_record(rec)
            if table:
                lines.append(f"- 表格文件：{rec.filename}\n{_spreadsheet_to_summary(table)}")
            else:
                # 解析失败（文件缺失/格式不支持/损坏）：明确告知模型，避免其按文件名幻觉编造表格内容
                lines.append(
                    f"- ⚠️ 表格文件：{rec.filename} 当前无法解析或文件不可达。"
                    f"严禁编造其内容，请直接告知用户该表格暂时无法读取，并请其重新上传。"
                )
            continue
        kind = "图片" if _is_image_record(rec) else "文件"
        lines.append(f"- {kind}：{rec.filename}（{rec.content_type or '未知类型'}，{rec.size} 字节）")
    return "\n".join(lines)


@router.post("", response_model=ChatResponse)
async def chat(
    request: ChatRequest,
    http_request: Request = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ChatResponse:
    """转发对话到 litellm 网关，支持 tool calling 循环执行 MES 操作。"""
    operator = current_user.username or current_user.id
    factory_id = (http_request.headers.get("x-factory-id") if http_request else None) or getattr(current_user, "active_factory_id", None) or current_user.factory_id or "FAC_MECH_001"

    last_user = next(
        (m.content for m in reversed(request.messages) if m.role == "user"), ""
    )
    actions: List[ToolAction] = []

    # ---- 加载本轮附件（工厂隔离）：图片走多模态 vision，非图片以文字摘要告知 ----
    att_records = await _load_attachment_records(db, request.attachments, current_user) \
        if request.attachments else []
    image_records = [r for r in att_records if _is_image_record(r)]

    task_id = MODEL_STACK_VISION_TASK_ID if image_records else MODEL_STACK_CHAT_TASK_ID
    prompt_tokens = max(1, sum(len(m.content or "") for m in request.messages) // 4)
    try:
        route = await _resolve_model_route(task_id, prompt_tokens=prompt_tokens)
    except Exception as exc:  # noqa: BLE001
        return ChatResponse(
            reply=_degraded_message(f"模型底座路由失败 ({type(exc).__name__})"),
            model=task_id, degraded=True, actions=actions,
        )

    # 所有文本意图统一交给模型解析；后端仅执行模型返回的 tool_calls。
    messages: List[Dict[str, Any]] = [{"role": "system", "content": SYSTEM_PROMPT}]
    # ---- 智能体调度：指定 agent 时注入其职责提示词，并记录监督心跳 ----
    if request.agent_key:
        agent_prompt = build_agent_system_prompt(request.agent_key)
        if agent_prompt:
            messages.append({"role": "system", "content": agent_prompt})
            await record_agent_dispatch(db, factory_id, request.agent_key, last_user)
    history = [m.model_dump() for m in request.messages]
    # 将附件注入「最后一条用户消息」：图片 → 多模态 content；非图片 → 文字摘要追加
    non_image_note = _attachment_text_note([r for r in att_records if not _is_image_record(r)])
    injected = False
    for idx in range(len(history) - 1, -1, -1):
        if history[idx].get("role") == "user":
            text = history[idx].get("content") or ""
            if non_image_note:
                text = f"{text}{non_image_note}"
            history[idx]["content"] = _build_multimodal_content(text, image_records)
            injected = True
            break
    if not injected and image_records:
        # 历史中无用户消息（异常场景）：补一条多模态用户消息
        history.append({"role": "user", "content": _build_multimodal_content(last_user, image_records)})
    messages += history

    payload: Dict[str, Any] = {
        "model": route["gateway_model"],
        "messages": messages,
        "temperature": request.temperature,
        "max_tokens": route["max_completion_tokens"],
    }
    # 图片内容由视觉任务模型理解；业务侧不再用 OCR 关键词选择具体模型。
    if not image_records and request.enable_tools:
        payload["tools"] = TOOL_DEFINITIONS
        payload["tool_choice"] = "auto"

    try:
        for round_index in range(MAX_TOOL_ROUNDS):
            resp = await _call_llm(
                payload,
                request_timeout=route["request_timeout"],
            )

            if resp.status_code >= 400:
                return ChatResponse(
                    reply=_degraded_message(f"网关返回 {resp.status_code}"),
                    model=route["task_id"], degraded=True, actions=actions,
                )

            data = resp.json()
            choice = data.get("choices", [{}])[0]
            message = choice.get("message", {}) or {}
            tool_calls = message.get("tool_calls") or []

            # 无工具调用 → 最终回复
            if not tool_calls:
                reply = _clean_model_reply(message.get("content") or "")
                if not reply:
                    return ChatResponse(
                        reply=_degraded_message("网关无有效回复"),
                        model=route["task_id"], degraded=True, actions=actions,
                    )
                reply = await _verify_grounded_reply(reply, actions, route)
                return ChatResponse(
                    reply=reply,
                    model=route["task_id"],
                    degraded=False,
                    actions=actions,
                )

            # 有工具调用 → 逐个执行，把 assistant 消息和 tool 结果追加到上下文
            messages.append({
                "role": "assistant",
                "content": message.get("content") or "",
                "tool_calls": tool_calls,
            })
            for tc in tool_calls:
                fn = tc.get("function", {}) or {}
                tool_name = fn.get("name", "")
                try:
                    arguments = json.loads(fn.get("arguments") or "{}")
                except (json.JSONDecodeError, TypeError):
                    arguments = {}

                result = await execute_tool(db, tool_name, arguments, operator=operator, factory_id=factory_id)
                is_error = "error" in result
                actions.append(ToolAction(
                    tool=tool_name,
                    label=TOOL_LABELS.get(tool_name, tool_name),
                    arguments=arguments,
                    result=result,
                    is_write=tool_name in WRITE_TOOLS,
                    is_sim=tool_name in SIM_TOOLS,
                    success=not is_error,
                ))
                messages.append({
                    "role": "tool",
                    "tool_call_id": tc.get("id", ""),
                    "content": _grounded_tool_result(result),
                })

            # 继续下一轮：工具清单保留，模型可基于已有结果继续取数或直接作答。
            messages.append({"role": "system", "content": FINAL_GROUNDING_PROMPT})
            payload["messages"] = messages
            if round_index + 1 >= MAX_TOOL_ROUNDS - 1:
                # 最后一轮不再给工具，强制模型收敛出面向用户的答复
                payload.pop("tools", None)
                payload.pop("tool_choice", None)

        # 超过最大轮次
        return ChatResponse(
            reply="操作轮次过多，已停止。请简化您的请求后重试。",
            model=route["task_id"], degraded=True, actions=actions,
        )
    except Exception as exc:  # noqa: BLE001
        return ChatResponse(
            reply=_degraded_message(f"网关连接失败 ({type(exc).__name__})"),
            model=route["task_id"], degraded=True, actions=actions,
        )


def _degraded_message(reason: str) -> str:
    return (
        f"AI 服务暂不可用（{reason}）。\n\n"
        "模型任务未能由模型底座正常下发，请检查控制面与模型网关状态。"
    )


# ==================== 结构化表格提取（chatbot → 在线表格） ====================

# 各查询工具结果 → 表格列定义（key=字段名, label=中文表头）
_TABLE_COLUMNS: Dict[str, List[Dict[str, str]]] = {
    "query_work_orders": [
        {"key": "work_order_code", "label": "工单号"},
        {"key": "product_name", "label": "产品"},
        {"key": "planned_qty", "label": "计划数"},
        {"key": "completed_qty", "label": "完成数"},
        {"key": "progress_pct", "label": "进度%"},
        {"key": "status", "label": "状态"},
        {"key": "priority", "label": "优先级"},
        {"key": "planned_due", "label": "交期"},
    ],
    "query_order_work_order_status": [
        {"key": "sales_order_code", "label": "销售订单"},
        {"key": "sales_order_status", "label": "订单状态"},
        {"key": "actual_work_order_count", "label": "生产工单数"},
        {"key": "master_work_order_code", "label": "主工单"},
        {"key": "master_status", "label": "主工单状态"},
        {"key": "operation_work_order_count", "label": "工序工单数"},
        {"key": "released_operation_count", "label": "已下达工序"},
        {"key": "pending_operation_count", "label": "待下达工序"},
        {"key": "all_operation_work_orders_released", "label": "工序是否全下达"},
    ],
    "query_inventory": [
        {"key": "material_code", "label": "物料编码"},
        {"key": "material_id", "label": "物料ID"},
        {"key": "warehouse_id", "label": "仓库"},
        {"key": "batch_code", "label": "批次"},
        {"key": "total_qty", "label": "总数量"},
        {"key": "available_qty", "label": "可用"},
        {"key": "reserved_qty", "label": "预留"},
        {"key": "status", "label": "状态"},
    ],
    "query_defects": [
        {"key": "record_code", "label": "记录编号"},
        {"key": "defect_type", "label": "不良类型"},
        {"key": "severity", "label": "严重度"},
        {"key": "quantity", "label": "数量"},
        {"key": "disposition", "label": "处置"},
        {"key": "root_cause_category", "label": "根因分类"},
        {"key": "description", "label": "描述"},
        {"key": "created_at", "label": "时间"},
    ],
    "query_equipment": [
        {"key": "equipment_code", "label": "设备编号"},
        {"key": "equipment_name", "label": "设备名称"},
        {"key": "equipment_type", "label": "类型"},
        {"key": "status", "label": "状态"},
        {"key": "station_id", "label": "工位"},
    ],
}

# 工具结果中存放列表数据的字段名
_TABLE_LIST_KEY: Dict[str, str] = {
    "query_work_orders": "work_orders",
    "query_order_work_order_status": "orders",
    "query_inventory": "inventory",
    "query_defects": "defects",
    "query_equipment": "equipment",
}

# 生产汇总 → 指标型表格（指标/数值 两列）
_SUMMARY_LABELS: Dict[str, str] = {
    "date": "日期",
    "today_good_output": "今日良品产出",
    "today_defect": "今日不良数",
    "yield_rate_pct": "良率(%)",
    "today_report_count": "今日报工次数",
    "active_work_orders": "在制工单",
    "pending_work_orders": "待开工单",
    "total_work_orders": "工单总数",
    "equipment_total": "设备总数",
    "equipment_running": "运行中设备",
    "equipment_fault": "故障设备",
    "equipment_utilization_pct": "设备利用率(%)",
}


def _extract_table_data(tool_name: str, result: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """从查询工具结果中提取结构化表格数据，供前端渲染可交互表格/电子表格。

    返回 {title, columns: [{key, label}], rows: [{...}]} 或 None（不适用表格的工具）。
    """
    # 指标汇总型 → 转为 指标/数值 两列表格
    if tool_name == "get_production_summary":
        rows = [
            {"metric": _SUMMARY_LABELS.get(k, k), "value": v}
            for k, v in result.items()
            if k != "error"
        ]
        if not rows:
            return None
        return {
            "title": "生产概况汇总",
            "columns": [{"key": "metric", "label": "指标"}, {"key": "value", "label": "数值"}],
            "rows": rows,
        }

    # 单条详情型 → 字段/值 两列表格
    if tool_name == "get_work_order_detail":
        label_map = {
            "work_order_code": "工单号", "product_name": "产品", "planned_qty": "计划数",
            "completed_qty": "完成数", "good_qty": "良品数", "defect_qty": "不良数",
            "scrap_qty": "报废数", "progress_pct": "进度(%)", "status": "状态",
            "priority": "优先级", "planned_due": "交期", "station_id": "工位",
            "routing_step": "当前工序", "created_at": "创建时间", "actual_start": "实际开工",
            "remark": "备注",
        }
        rows = [
            {"field": label_map.get(k, k), "value": v}
            for k, v in result.items()
            if k not in ("id", "product_id", "error") and v is not None
        ]
        if not rows:
            return None
        return {
            "title": f"工单详情 {result.get('work_order_code', '')}",
            "columns": [{"key": "field", "label": "字段"}, {"key": "value", "label": "值"}],
            "rows": rows,
        }

    # 列表型查询工具
    list_key = _TABLE_LIST_KEY.get(tool_name)
    columns = _TABLE_COLUMNS.get(tool_name)
    if not list_key or not columns:
        # ---- 流程知识工具：根据返回类型动态构建表格 ----
        if tool_name == "query_process_knowledge":
            return _extract_knowledge_table(result)
        return None
    items = result.get(list_key)
    if not isinstance(items, list) or not items:
        return None
    return {
        "title": TOOL_LABELS.get(tool_name, tool_name),
        "columns": columns,
        "rows": items,
    }


def _extract_knowledge_table(result: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """流程知识查询结果 → 结构化表格（工单流/职位SOP/RACI/概览）。"""
    rtype = result.get("type", "")

    # 工单全生命周期 / 单阶段
    if rtype in ("work_order_flow", "work_order_stage"):
        stages = result.get("stages") or []
        if not stages:
            return None
        return {
            "title": result.get("title", "工单流程"),
            "columns": [
                {"key": "stage", "label": "阶段"},
                {"key": "status", "label": "状态"},
                {"key": "role", "label": "负责角色"},
                {"key": "actions", "label": "动作"},
                {"key": "blockpoint", "label": "卡点/异常处理"},
            ],
            "rows": stages,
        }

    # 职位 SOP → 步骤表
    if rtype == "position_sop":
        pos = result.get("position") or {}
        flow = pos.get("daily_flow") or []
        if not flow:
            return None
        rows = [
            {"step": s.get("step"), "task": s.get("task"), "detail": s.get("detail")}
            for s in flow
        ]
        return {
            "title": result.get("title", "职位SOP"),
            "columns": [
                {"key": "step", "label": "序号"},
                {"key": "task", "label": "任务"},
                {"key": "detail", "label": "说明"},
            ],
            "rows": rows,
        }

    # RACI 责任归属
    if rtype == "who_handles":
        raci = result.get("raci") or []
        if not raci:
            return None
        return {
            "title": result.get("title", "责任归属"),
            "columns": [
                {"key": "role", "label": "角色"},
                {"key": "responsibility", "label": "RACI"},
                {"key": "meaning", "label": "含义"},
            ],
            "rows": raci,
        }

    # 全阶段主要负责人
    if rtype == "who_handles_all":
        stages = result.get("stages") or []
        if not stages:
            return None
        return {
            "title": result.get("title", "各环节负责人"),
            "columns": [
                {"key": "stage", "label": "阶段"},
                {"key": "status", "label": "状态"},
                {"key": "primary_role", "label": "主要负责人"},
                {"key": "blockpoint", "label": "卡点处理"},
            ],
            "rows": stages,
        }

    # 职位概览
    if rtype == "position_overview":
        positions = result.get("positions") or []
        if not positions:
            return None
        return {
            "title": result.get("title", "职位概览"),
            "columns": [
                {"key": "position", "label": "职位"},
                {"key": "duties", "label": "核心职责"},
                {"key": "steps", "label": "SOP步骤数"},
            ],
            "rows": positions,
        }

    return None


# ==================== 流式输出（SSE） ====================

def _sse(event: str, data: Any) -> str:
    """格式化一条 SSE 帧。"""
    return f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False, default=str)}\n\n"


async def _stream_llm_deltas(
    payload: Dict[str, Any],
    *,
    request_timeout: float,
):
    """调用网关 SSE，逐块返回 OpenAI delta。"""
    headers = {
        "Content-Type": "application/json",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
    }
    if API_KEY:
        headers["Authorization"] = f"Bearer {API_KEY}"
    stream_payload = {
        **payload,
        "stream": True,
        "cache": {"no-cache": True},
    }
    async with httpx.AsyncClient(timeout=request_timeout) as client:
        async with client.stream(
            "POST",
            f"{GATEWAY_URL}/v1/chat/completions",
            json=stream_payload,
            headers=headers,
        ) as resp:
            if resp.status_code >= 400:
                body = await resp.aread()
                raise RuntimeError(
                    f"gateway {resp.status_code}: {body[:300].decode(errors='replace')}"
                )
            async for line in resp.aiter_lines():
                if not line.startswith("data: "):
                    continue
                data_str = line[6:].strip()
                if data_str == "[DONE]":
                    break
                try:
                    chunk = json.loads(data_str)
                    yield chunk.get("choices", [{}])[0].get("delta", {}) or {}
                except (json.JSONDecodeError, IndexError, KeyError):
                    continue


def _merge_stream_tool_calls(
    accumulated: Dict[int, Dict[str, Any]],
    deltas: List[Dict[str, Any]],
) -> None:
    """合并 OpenAI 流式 tool_calls 的分片。"""
    for item in deltas:
        index = int(item.get("index", 0))
        target = accumulated.setdefault(index, {
            "id": "",
            "type": "function",
            "function": {"name": "", "arguments": ""},
        })
        if item.get("id"):
            target["id"] = item["id"]
        if item.get("type"):
            target["type"] = item["type"]
        function_delta = item.get("function") or {}
        if function_delta.get("name"):
            target["function"]["name"] += function_delta["name"]
        if function_delta.get("arguments"):
            target["function"]["arguments"] += function_delta["arguments"]


@router.post("/stream")
async def chat_stream(
    request: ChatRequest,
    http_request: Request = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """SSE 流式对话：工具执行实时推送 action 事件，最终回复逐 token 流式输出。"""
    operator = current_user.username or current_user.id
    factory_id = (http_request.headers.get("x-factory-id") if http_request else None) or getattr(current_user, "active_factory_id", None) or current_user.factory_id or "FAC_MECH_001"

    last_user = next(
        (m.content for m in reversed(request.messages) if m.role == "user"), ""
    )

    async def generate():
        actions: List[ToolAction] = []

        # ---- 加载附件 ----
        att_records = await _load_attachment_records(db, request.attachments, current_user) \
            if request.attachments else []
        image_records = [r for r in att_records if _is_image_record(r)]

        task_id = MODEL_STACK_VISION_TASK_ID if image_records else MODEL_STACK_CHAT_TASK_ID
        prompt_tokens = max(1, sum(len(m.content or "") for m in request.messages) // 4)
        try:
            route = await _resolve_model_route(task_id, prompt_tokens=prompt_tokens)
        except Exception as exc:  # noqa: BLE001
            yield _sse("delta", {
                "content": _degraded_message(
                    f"模型底座路由失败 ({type(exc).__name__})"
                ),
            })
            yield _sse("done", {"model": task_id, "degraded": True})
            return

        # ---- Excel/CSV 附件 → 推送结构化表格事件（前端渲染可交互表格 + Univer 电子表格） ----
        for rec in att_records:
            if _is_spreadsheet_record(rec):
                tbl = _parse_spreadsheet_record(rec)
                if tbl:
                    yield _sse("table", tbl)

        # 所有文本意图统一交给模型解析；后端仅执行模型返回的 tool_calls。
        messages: List[Dict[str, Any]] = [{"role": "system", "content": SYSTEM_PROMPT}]
        # ---- 智能体调度：指定 agent 时注入其职责提示词，并记录监督心跳 ----
        if request.agent_key:
            agent_prompt = build_agent_system_prompt(request.agent_key)
            if agent_prompt:
                messages.append({"role": "system", "content": agent_prompt})
                await record_agent_dispatch(db, factory_id, request.agent_key, last_user)
        history = [m.model_dump() for m in request.messages]
        non_image_note = _attachment_text_note([r for r in att_records if not _is_image_record(r)])
        injected = False
        for idx in range(len(history) - 1, -1, -1):
            if history[idx].get("role") == "user":
                text = history[idx].get("content") or ""
                if non_image_note:
                    text = f"{text}{non_image_note}"
                history[idx]["content"] = _build_multimodal_content(text, image_records)
                injected = True
                break
        if not injected and image_records:
            history.append({"role": "user", "content": _build_multimodal_content(last_user, image_records)})
        messages += history

        payload: Dict[str, Any] = {
            "model": route["gateway_model"],
            "messages": messages,
            "temperature": request.temperature,
            "max_tokens": route["max_completion_tokens"],
        }
        # 图片内容由视觉任务模型做语义理解，不在业务侧做关键词分流。
        if not image_records and request.enable_tools:
            payload["tools"] = TOOL_DEFINITIONS
            payload["tool_choice"] = "auto"

        try:
            for round_index in range(MAX_TOOL_ROUNDS):
                streamed_content: List[str] = []
                streamed_tool_calls: Dict[int, Dict[str, Any]] = {}
                async for delta in _stream_llm_deltas(
                    payload,
                    request_timeout=route["request_timeout"],
                ):
                    text = delta.get("content") or ""
                    if text:
                        streamed_content.append(text)
                        yield _sse("delta", {"content": text})
                    _merge_stream_tool_calls(
                        streamed_tool_calls,
                        delta.get("tool_calls") or [],
                    )

                tool_calls = [
                    streamed_tool_calls[index]
                    for index in sorted(streamed_tool_calls)
                    if streamed_tool_calls[index]["function"]["name"]
                ]
                if not tool_calls:
                    if not streamed_content:
                        yield _sse("delta", {"content": _degraded_message("网关无有效回复")})
                        yield _sse("done", {"model": route["task_id"], "degraded": True})
                        return
                    yield _sse("done", {"model": route["task_id"], "degraded": False})
                    return

                # 模型通过流协议下发工具调用，后端执行并实时推送 action。
                messages.append({
                    "role": "assistant",
                    "content": "".join(streamed_content),
                    "tool_calls": tool_calls,
                })
                for tc in tool_calls:
                    fn = tc.get("function", {}) or {}
                    tool_name = fn.get("name", "")
                    try:
                        arguments = json.loads(fn.get("arguments") or "{}")
                    except (json.JSONDecodeError, TypeError):
                        arguments = {}
                    result = await execute_tool(db, tool_name, arguments, operator=operator, factory_id=factory_id)
                    is_error = "error" in result
                    action = ToolAction(
                        tool=tool_name,
                        label=TOOL_LABELS.get(tool_name, tool_name),
                        arguments=arguments,
                        result=result,
                        is_write=tool_name in WRITE_TOOLS,
                        is_sim=tool_name in SIM_TOOLS,
                        success=not is_error,
                    )
                    actions.append(action)
                    yield _sse("action", action.model_dump())
                    table_data = _extract_table_data(tool_name, result)
                    if table_data:
                        yield _sse("table", table_data)
                    messages.append({
                        "role": "tool",
                        "tool_call_id": tc.get("id", ""),
                        "content": _grounded_tool_result(result),
                    })

                # 继续下一轮：工具清单保留，模型可基于已有结果继续取数或直接作答。
                messages.append({
                    "role": "system",
                    "content": (
                        f"{FINAL_GROUNDING_PROMPT}"
                        "请基于上述工具结果，用简洁自然的中文回答用户问题，"
                        "只输出面向用户的最终答复。"
                    ),
                })
                payload["messages"] = messages
                if round_index + 1 >= MAX_TOOL_ROUNDS - 1:
                    # 最后一轮不再给工具，强制模型收敛出面向用户的答复
                    payload.pop("tools", None)
                    payload.pop("tool_choice", None)

            yield _sse("delta", {"content": "操作轮次过多，已停止。请简化您的请求后重试。"})
            yield _sse("done", {"model": route["task_id"], "degraded": True})
        except Exception as exc:  # noqa: BLE001
            yield _sse("delta", {"content": _degraded_message(f"网关连接失败 ({type(exc).__name__})")})
            yield _sse("done", {"model": route["task_id"], "degraded": True})

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


__all__ = ["router"]
